"""
backend/routes/excel_io.py
Shared helpers for Excel import and export across all StockFlow modules.
"""
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import send_file

# ── Style constants ───────────────────────────────────────────────────────────
_H_FILL  = PatternFill("solid", start_color="1E3A5F", end_color="1E3A5F")
_H_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
_B_FONT  = Font(name="Arial", size=10)
_BORDER  = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)
_NOTE_FILL = PatternFill("solid", start_color="FFF9C4", end_color="FFF9C4")
_NOTE_FONT = Font(italic=True, color="7B6000", name="Arial", size=9)


def _header_row(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = _H_FILL
        c.font = _H_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = _BORDER
    ws.row_dimensions[row].height = 20


def _body_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _B_FONT
    c.border = _BORDER
    c.alignment = Alignment(vertical='center', wrap_text=True)
    return c


def _col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _stream(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


def _read_upload(file_storage):
    """Return list-of-dicts from an uploaded .xlsx file (row 1 = headers)."""
    wb = openpyxl.load_workbook(file_storage, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h else '' for h in rows[0]]
    records = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        records.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
    return records


# ═══════════════════════════════════════════════════════════════════════════════
# PRODUCTS
# ═══════════════════════════════════════════════════════════════════════════════

PRODUCT_COLS = ["name","description","unit","purchase_price","selling_price",
                "current_quantity","min_stock","max_stock"]
PRODUCT_WIDTHS = [22, 25, 10, 20, 20, 18, 12, 12]


def export_products(products):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    _header_row(ws, PRODUCT_COLS)
    for r, p in enumerate(products, 2):
        _body_cell(ws, r, 1, p.get('name',''))
        _body_cell(ws, r, 2, p.get('description',''))
        _body_cell(ws, r, 3, p.get('unit',''))
        _body_cell(ws, r, 4, float(p.get('purchase_price', 0)))
        _body_cell(ws, r, 5, float(p.get('selling_price', 0)))
        _body_cell(ws, r, 6, int(p.get('current_quantity', 0)))
        _body_cell(ws, r, 7, int(p.get('min_stock', 0)))
        _body_cell(ws, r, 8, p.get('max_stock', ''))
    _col_widths(ws, PRODUCT_WIDTHS)
    ts = datetime.utcnow().strftime('%Y%m%d')
    return _stream(wb, f"products_export_{ts}.xlsx")


def import_products(file_storage, db, business_id_obj):
    """
    Returns: (inserted, updated, duplicates, errors)
    duplicates = list of {name, existing_id} that need user decision
    """
    records = _read_upload(file_storage)
    inserted, updated, duplicates, errors = [], [], [], []

    for i, row in enumerate(records, 2):
        name = str(row.get('name') or '').strip()
        if not name:
            errors.append(f"Row {i}: name is required")
            continue

        try:
            data = {
                "name": name,
                "description": str(row.get('description') or '').strip(),
                "unit": str(row.get('unit') or 'piece').strip(),
                "purchase_price": float(row.get('purchase_price') or 0),
                "selling_price":  float(row.get('selling_price') or 0),
                "current_quantity": int(float(row.get('current_quantity') or 0)),
                "min_stock": int(float(row.get('min_stock') or 0)),
                "max_stock": int(float(row.get('max_stock'))) if row.get('max_stock') else None,
            }
        except (ValueError, TypeError) as e:
            errors.append(f"Row {i} ({name}): invalid number — {e}")
            continue

        query = {"name": {"$regex": f"^{name}$", "$options": "i"}}
        if business_id_obj:
            query["business_id"] = business_id_obj

        existing = db.products.find_one(query)
        if existing:
            duplicates.append({"row": i, "name": name, "existing_id": str(existing['_id']), "data": data})
        else:
            if business_id_obj:
                data["business_id"] = business_id_obj
            data["created_at"] = datetime.utcnow()
            db.products.insert_one(data)
            inserted.append(name)

    return inserted, updated, duplicates, errors


def resolve_product_duplicate(db, existing_id, data, business_id_obj, action):
    """action: 'overwrite' or 'skip'"""
    if action == 'overwrite':
        from bson.objectid import ObjectId
        update_data = {k: v for k, v in data.items() if k not in ('business_id', 'created_at')}
        db.products.update_one({"_id": ObjectId(existing_id)}, {"$set": update_data})
        return True
    return False


# ═══════════════════════════════════════════════════════════════════════════════
# SUPPLIERS
# ═══════════════════════════════════════════════════════════════════════════════

SUPPLIER_COLS = ["name","contact_person","phone","email","address"]
SUPPLIER_WIDTHS = [25, 22, 16, 28, 35]


def export_suppliers(suppliers):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suppliers"
    _header_row(ws, SUPPLIER_COLS)
    for r, s in enumerate(suppliers, 2):
        _body_cell(ws, r, 1, s.get('name',''))
        _body_cell(ws, r, 2, s.get('contact_person',''))
        _body_cell(ws, r, 3, s.get('phone',''))
        _body_cell(ws, r, 4, s.get('email',''))
        _body_cell(ws, r, 5, s.get('address',''))
    _col_widths(ws, SUPPLIER_WIDTHS)
    ts = datetime.utcnow().strftime('%Y%m%d')
    return _stream(wb, f"suppliers_export_{ts}.xlsx")


def import_suppliers(file_storage, db, business_id_obj):
    records = _read_upload(file_storage)
    inserted, duplicates, errors = [], [], []

    for i, row in enumerate(records, 2):
        name = str(row.get('name') or '').strip()
        if not name:
            errors.append(f"Row {i}: name is required")
            continue

        data = {
            "name": name,
            "contact_person": str(row.get('contact_person') or '').strip(),
            "phone": str(row.get('phone') or '').strip(),
            "email": str(row.get('email') or '').strip(),
            "address": str(row.get('address') or '').strip(),
        }

        query = {"name": {"$regex": f"^{name}$", "$options": "i"}}
        if business_id_obj:
            query["business_id"] = business_id_obj

        existing = db.suppliers.find_one(query)
        if existing:
            duplicates.append({"row": i, "name": name, "existing_id": str(existing['_id']), "data": data})
        else:
            if business_id_obj:
                data["business_id"] = business_id_obj
            data["created_at"] = datetime.utcnow()
            db.suppliers.insert_one(data)
            inserted.append(name)

    return inserted, duplicates, errors


def resolve_supplier_duplicate(db, existing_id, data, action):
    if action == 'overwrite':
        from bson.objectid import ObjectId
        db.suppliers.update_one({"_id": ObjectId(existing_id)}, {"$set": data})
        return True
    return False


# ═══════════════════════════════════════════════════════════════════════════════
# CLIENTS
# ═══════════════════════════════════════════════════════════════════════════════

CLIENT_COLS = ["name","contact","kra_pin"]
CLIENT_WIDTHS = [25, 18, 18]


def export_clients(clients):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients"
    _header_row(ws, CLIENT_COLS + ["balance"])
    for r, c in enumerate(clients, 2):
        _body_cell(ws, r, 1, c.get('name',''))
        _body_cell(ws, r, 2, c.get('contact',''))
        _body_cell(ws, r, 3, c.get('kra_pin',''))
        _body_cell(ws, r, 4, float(c.get('balance', 0)))
    _col_widths(ws, CLIENT_WIDTHS + [16])
    ts = datetime.utcnow().strftime('%Y%m%d')
    return _stream(wb, f"clients_export_{ts}.xlsx")


def import_clients(file_storage, db, business_id_obj):
    records = _read_upload(file_storage)
    inserted, duplicates, errors = [], [], []

    for i, row in enumerate(records, 2):
        name = str(row.get('name') or '').strip()
        if not name:
            errors.append(f"Row {i}: name is required")
            continue

        data = {
            "name": name,
            "contact": str(row.get('contact') or '').strip(),
            "kra_pin": str(row.get('kra_pin') or '').strip(),
        }

        query = {"name": {"$regex": f"^{name}$", "$options": "i"}}
        if business_id_obj:
            query["business_id"] = business_id_obj

        existing = db.clients.find_one(query)
        if existing:
            duplicates.append({"row": i, "name": name, "existing_id": str(existing['_id']), "data": data})
        else:
            if business_id_obj:
                data["business_id"] = business_id_obj
            data["balance"] = 0.0
            data["created_at"] = datetime.utcnow()
            db.clients.insert_one(data)
            inserted.append(name)

    return inserted, duplicates, errors


def resolve_client_duplicate(db, existing_id, data, action):
    """Only update contact/kra_pin — never touch balance."""
    if action == 'overwrite':
        from bson.objectid import ObjectId
        safe = {k: v for k, v in data.items() if k not in ('balance', 'business_id', 'created_at')}
        db.clients.update_one({"_id": ObjectId(existing_id)}, {"$set": safe})
        return True
    return False


# ═══════════════════════════════════════════════════════════════════════════════
# PURCHASES — export only
# ═══════════════════════════════════════════════════════════════════════════════

def export_purchases(purchases):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchases"
    headers = ["date","supplier","product","quantity","cost_price","line_total","purchase_total"]
    _header_row(ws, headers)
    r = 2
    for p in purchases:
        date_str = p['date'].strftime('%Y-%m-%d %H:%M') if p.get('date') else ''
        supplier = p.get('supplier_name', '')
        total = float(p.get('total_cost', 0))
        first = True
        for item in p.get('items', []):
            qty   = float(item.get('quantity', 0))
            price = float(item.get('cost_price', 0))
            _body_cell(ws, r, 1, date_str if first else '')
            _body_cell(ws, r, 2, supplier if first else '')
            _body_cell(ws, r, 3, item.get('product_name', ''))
            _body_cell(ws, r, 4, qty)
            _body_cell(ws, r, 5, price)
            _body_cell(ws, r, 6, qty * price)
            _body_cell(ws, r, 7, total if first else '')
            first = False
            r += 1
        if not p.get('items'):
            _body_cell(ws, r, 1, date_str)
            _body_cell(ws, r, 2, supplier)
            for col in range(3, 7): _body_cell(ws, r, col, '')
            _body_cell(ws, r, 7, total)
            r += 1
    _col_widths(ws, [18, 22, 22, 10, 14, 14, 16])
    ts = datetime.utcnow().strftime('%Y%m%d')
    return _stream(wb, f"purchases_export_{ts}.xlsx")


# ═══════════════════════════════════════════════════════════════════════════════
# SALES — export only
# ═══════════════════════════════════════════════════════════════════════════════

def export_sales(sales):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    headers = ["date","cashier","product","quantity","selling_price","line_total","sale_total","payment_method"]
    _header_row(ws, headers)
    r = 2
    for s in sales:
        date_str = s['date'].strftime('%Y-%m-%d %H:%M') if s.get('date') else ''
        cashier  = s.get('cashier_name', '')
        total    = float(s.get('total_amount', 0))
        method   = s.get('payment_method', '')
        first = True
        for item in s.get('items', []):
            qty   = float(item.get('quantity', 0))
            price = float(item.get('selling_price', 0))
            _body_cell(ws, r, 1, date_str if first else '')
            _body_cell(ws, r, 2, cashier if first else '')
            _body_cell(ws, r, 3, item.get('product_name', ''))
            _body_cell(ws, r, 4, qty)
            _body_cell(ws, r, 5, price)
            _body_cell(ws, r, 6, qty * price)
            _body_cell(ws, r, 7, total if first else '')
            _body_cell(ws, r, 8, method if first else '')
            first = False
            r += 1
        if not s.get('items'):
            _body_cell(ws, r, 1, date_str)
            _body_cell(ws, r, 2, cashier)
            for col in range(3, 8): _body_cell(ws, r, col, '')
            _body_cell(ws, r, 7, total)
            _body_cell(ws, r, 8, method)
            r += 1
    _col_widths(ws, [18, 18, 22, 10, 14, 14, 14, 16])
    ts = datetime.utcnow().strftime('%Y%m%d')
    return _stream(wb, f"sales_export_{ts}.xlsx")